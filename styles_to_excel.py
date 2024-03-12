# -*- coding: utf-8 -*-
"""
Created on Thu Sep 15 14:47:27 2022

@author: zarub
"""



from openpyxl import load_workbook
from openpyxl.styles import  Side,  Alignment
from openpyxl.styles import  PatternFill, Font, Border

# import xlsxwriter

import logging
rootLogger = logging.getLogger(__name__)

def excel_style(path,root):
    
    light_purple = "00CC99FF"
    green = "00008000"
    yellow = "FFFF30"
    blue = "00008B"
    black = '000000'
    gray = 'F0F0F8'
    thin = Side(border_style="thin", color=light_purple)
    double = Side(border_style="double", color=green)
    
    wb = load_workbook(path)
    wb.guess_types = True
    sheets = wb.sheetnames

    
    for j in range(0,len(sheets)):
        sheet = wb[sheets[j]]
        sheet_name = str(sheet)
        for i in range (1,sheet.max_column+1):
            
            cell = sheet.cell(row=1,column=i)
            col=xlsxwriter.utility.xl_col_to_name(i)
            sheet.column_dimensions[col].bestFit = True
            # sheet.column_dimensions[col].auto_size = True
            if sheets[j] == 'Globals':
                sheet.column_dimensions['A'].width = 50
            elif sheets[j] == 'pz_Data':
                sheet.column_dimensions['A'].width = 14
                sheet.column_dimensions['D'].width = 4
                sheet.column_dimensions['G'].width = 4
            elif sheets[j] == 'Calculated':
                sheet.column_dimensions['A'].width = 14
                sheet.column_dimensions['H'].width = 18
                sheet.column_dimensions['I'].width = 19
                sheet.column_dimensions['Y'].width = 3
            elif sheet_name.count('Group'):
                sheet.column_dimensions['A'].width = 14
                sheet.column_dimensions['I'].width = 18
                sheet.column_dimensions['J'].width = 19
                sheet.column_dimensions['Q'].width = 3
            else: 
                pass
                # sheet.column_dimensions[col].auto_size = True
                # sheet.column_dimensions[col].bestFit = True
            if sheets[j] == 'pz_Data':
                sheet.guess_types = True
                sheet['H8'].number_format ='0.000E+00'
                
                
            cell.border = Border(top=double, bottom=thin, left=thin, right=thin )
            cell.fill = PatternFill(start_color=yellow,  end_color=yellow,   fill_type='solid')
            cell.font = Font(b=True, color="FF0000", size=11)
            cell.alignment = Alignment(horizontal="center",
                                            vertical="center")
            
        for i in range (1,sheet.max_column+1):
            if sheets[j] != 'Globals':
                cell = sheet.cell(row=2,column=i) 
                cell.border = Border( left=thin, right=thin,
                                              bottom=double)
                cell.fill = PatternFill(start_color=gray,  end_color=gray,   fill_type='solid')
                cell.font = Font(b=False, color="060f76", size=8)
                cell.alignment = Alignment(horizontal="center",
                                            vertical="center")   
    wb.save(path)
    
    return path
    
# if __name__ == "__main__":
#     LogFilePath = 'D:/Phyton_MBE_exe/Project_Shablons/Out_files/Shablon_Kobz_3_grwell_out_10.xlsx'
#     root = tk.Tk()
#     path = excel_style(LogFilePath,root)
#     print(path)
#     # root.mainloop()
#     root.destroy()